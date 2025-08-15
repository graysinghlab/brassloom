
"""
brassloom_sync_gsu.py
- Reads opportunities.json and pushes selected/new items into GSU_Cayuse_Lite.xlsx
- Creates standard pre-award tasks and sets internal deadlines (configurable)
Usage:
  python brassloom_sync_gsu.py --ops opportunities.json --wb GSU_Cayuse_Lite.xlsx --all
  python brassloom_sync_gsu.py --ops opportunities.json --wb GSU_Cayuse_Lite.xlsx --filter "HBCU,MSI"
"""
import argparse, datetime, json, os, re
from typing import List, Dict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import yaml

CONFIG_PATH = os.path.join(os.path.dirname(__file__), "brassloom_config.yaml")

def load_config():
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)

def parse_date(s: str):
    if not s:
        return None
    s = s[:10]
    for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%Y/%m/%d"]:
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

def next_id(existing_ids: List[str], prefix="GSU-P-"):
    nums = []
    for i in existing_ids:
        m = re.match(rf"{re.escape(prefix)}(\d+)", str(i))
        if m: 
            try: nums.append(int(m.group(1)))
            except: pass
    nxt = (max(nums) + 1) if nums else 1
    return f"{prefix}{nxt:04d}"

def next_task_id(existing_ids: List[str]):
    nums = []
    for i in existing_ids:
        m = re.match(r"TSK-(\d+)", str(i))
        if m:
            try: nums.append(int(m.group(1)))
            except: pass
    nxt = (max(nums) + 1) if nums else 1
    return lambda : f"TSK-{(nums.append(nums[-1]+1) or nums[-1]) if nums else (nums.append(1) or 1):04d}"

def sponsor_type_from_agency(agency: str) -> str:
    if not agency: return ""
    agency_low = agency.lower()
    if any(k in agency_low for k in ["national science foundation","nih","health","grants.gov","nasa","nsf","department of","dod","doe","usda","epa","nsf funding","nih guide"]):
        return "Federal"
    if "board of regents" in agency_low or "state" in agency_low:
        return "State"
    return "Nonprofit"

def mechanism_from_source(item: Dict, cfg: Dict) -> str:
    src = (item.get("source") or item.get("agency") or "").strip()
    for k, v in cfg.get("mechanism_map", {}).items():
        if k.lower() in src.lower():
            return v
    # Grants.gov link heuristic
    if "grants.gov" in (item.get("url","")).lower(): return "Grants.gov"
    return "Other"

def prepare_rows(item: Dict, cfg: Dict, next_pid: str):
    title = item.get("title","").strip()
    sponsor = (item.get("agency") or item.get("source") or "").strip()
    sponsor_type = sponsor_type_from_agency(sponsor)
    close_date = parse_date(item.get("close_date",""))
    posted = parse_date(item.get("posted_date",""))

    due = close_date or posted
    internal = due - datetime.timedelta(days=cfg.get("internal_deadline_offset_days", 7)) if due else None

    props_row = [
        next_pid,                          # ProposalID
        title,                             # Title
        cfg["default_pi"]["id"],           # PI_ID
        cfg["default_pi"]["name"],         # PI_Name
        cfg["default_pi"]["dept"],         # Department
        cfg["default_pi"]["college"],      # College/Unit
        "",                                # SponsorID
        sponsor,                           # SponsorName
        sponsor_type,                      # SponsorType
        (item.get("id") or item.get("assistance_listing") or ""),  # FundingOpportunity
        internal.isoformat() if internal else "",                  # InternalDeadline
        due.isoformat() if due else "",                            # DueDate
        mechanism_from_source(item, cfg),  # SubmissionMechanism
        cfg.get("default_proposal_type","New"),                    # ProposalType
        cfg.get("default_status","Department Review"),             # Status
        "", "", "",                        # PrimeSponsorID, ProjectStart, ProjectEnd
        "", "", "",                        # TotalDirect, TotalIndirect, F&A_Rate
        "No", "", "",                      # CostShareRequired, Amount, Approvers
        "No","No","No","Yes","No","No",    # Human, Animal, Biosafety, COI, Export, DataSec
        0,                                 # SubawardsCount
        f"Imported by BrassLoom on {datetime.date.today().isoformat()}"
    ]

    mechanism = props_row[12]
    due_for_tasks = due or (datetime.date.today() + datetime.timedelta(days=30))
    def t(n): return (due_for_tasks - datetime.timedelta(days=n)).isoformat()

    tasks = [
        ("Complete GSU Internal Routing Form", t(10), "PI", "Pending", "Attach signed PDF"),
        ("COI Disclosures for all key personnel", t(9), "PI/OSP", "Pending", ""),
        ("Subrecipient Commitment Form(s)", t(8), "OSP", "Pending", "Collect UEI and F&A rate docs"),
        ("Export Control & Data Security review", t(8), "Compliance", "Pending", "If foreign collaborators or controlled data"),
        ("Final Budget & Justification", t(7), "OSP Pre-Award", "Pending", "Check salary cap and F&A base"),
        (f"Create application in {mechanism}", t(7), "OSP Pre-Award", "Pending", "Confirm FOA & forms"),
        ("Dean/Provost cost-share letter (if required)", t(7), "Dean/Provost", "Pending", "Upload letter")
    ]
    return props_row, tasks

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--ops", default=os.path.join(os.path.dirname(__file__), "opportunities.json"))
    ap.add_argument("--wb", default=os.path.join(os.path.dirname(__file__), "GSU_Cayuse_Lite.xlsx"))
    ap.add_argument("--filter", default="")  # comma-separated keywords to restrict import
    ap.add_argument("--all", action="store_true")
    args = ap.parse_args()

    cfg = load_config()

    with open(args.ops, "r", encoding="utf-8") as f:
        ops = json.load(f)

    if not os.path.exists(args.wb):
        raise SystemExit(f"Workbook not found: {args.wb}")

    wb = load_workbook(args.wb)
    wsP = wb["Proposals"]
    wsT = wb["Tasks"]

    existing_titles = set()
    existing_ids = []
    for row in wsP.iter_rows(min_row=2, values_only=True):
        if not any(row): continue
        existing_ids.append(str(row[0]))
        if row[1]: existing_titles.add(str(row[1]).strip().lower())

    next_pid = next_id(existing_ids, "GSU-P-")

    task_ids = [r[0] for r in wsT.iter_rows(min_row=2, values_only=True) if r and r[0]]
    next_task = next_task_id(task_ids)

    chosen = []
    if args.all:
        chosen = ops
    else:
        flt = [s.strip().lower() for s in args.filter.split(",") if s.strip()]
        if not flt:
            # default: import only items containing HBCU/MSI keywords
            flt = [k.lower() for k in cfg["keywords"]]
        for o in ops:
            blob = json.dumps(o).lower()
            if any(k in blob for k in flt):
                chosen.append(o)

    added = 0
    for item in chosen:
        title = (item.get("title") or "").strip().lower()
        if not title or title in existing_titles:
            continue
        prop_row, tasks = prepare_rows(item, cfg, next_pid)
        wsP.append(prop_row)

        for name, due, owner, status, notes in tasks:
            wsT.append([next_task(), prop_row[0], name, due, owner, status, notes])

        existing_titles.add(title)
        existing_ids.append(next_pid)
        # increment proposal id for next one
        m = re.match(r"GSU-P-(\d+)", next_pid)
        nxt = int(m.group(1)) + 1 if m else 1
        next_pid = f"GSU-P-{nxt:04d}"
        added += 1

    wb.save(args.wb)
    print(f"Imported {added} opportunities into {os.path.basename(args.wb)}")

if __name__ == "__main__":
    main()
